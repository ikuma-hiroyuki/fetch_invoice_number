import os
import subprocess
from dataclasses import dataclass
from pathlib import Path

import openpyxl


@dataclass
class Columns:
    """ 登録番号シートの列名を管理するクラス """
    corporate_name = 1
    invoice_number = 2
    corporate_number = 3
    tax_agency_registered_name = 4
    address = 5


class InvoiceExcelProcessor:
    """ インボイス情報取得対象企業のExcelファイルの読み書きを行うクラス """
    base_dir = Path(__file__).resolve().parent
    csv_dir = base_dir / "csv"
    target_excel = csv_dir / "sample_excel_copy.xlsx"

    def __init__(self):
        self.target_corporations = self._create_target_corporations()

    @classmethod
    def is_open_output_excel(cls) -> bool:
        """
        excel_pathが開かれているかどうかを返す
        :return: excel_pathが開かれているかどうか
        """
        if os.name == "nt":
            # Windows
            try:
                with cls.target_excel.open("r+b"):
                    pass
                return False
            except IOError:
                return True
        elif os.name == "posix":
            # Unix
            if cls.target_excel.exists():
                result = subprocess.run(["lsof", str(cls.target_excel)], stdout=subprocess.PIPE)
                return bool(result.stdout)
            else:
                return False

    @staticmethod
    def generate_corporate_number(invoice_number: str) -> str:
        """
        登録番号から法人番号を生成する
        :param invoice_number: 登録番号
        :return: 法人番号
        """
        return invoice_number.replace("-", "")[1:]

    def _create_target_corporations(self) -> list[dict[str, str, str]]:
        """
        csv/target.csvから登録番号、事業者名、法人番号の辞書を内包するリスト作成し返す

        A列に事業者名、B列に登録番号(Tから始まる13桁の番号)。1行目はヘッダー行。
        :return: 登録番号、事業者名、法人番号の辞書を内包するリスト
        """

        try:
            wb = openpyxl.load_workbook(self.target_excel)
            ws = wb["登録番号"]
            target_corporations = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                invoice_number = ws.cell(row_num, Columns.invoice_number).value
                target_corporations.append({
                    "登録番号": invoice_number,
                    "事業者名": ws.cell(row_num, Columns.corporate_name).value,
                    "法人番号": self.generate_corporate_number(invoice_number)
                })
            wb.close()
        except FileNotFoundError:
            target_corporations = []
        return target_corporations

    def write_result_excel(self, fetched_data: dict[str, dict[str, str]]):
        """
        取込結果をExcelに書き込む
        :param fetched_data: APIから取得した法人番号と法人名、住所の辞書
        """

        wb = openpyxl.load_workbook(self.target_excel)
        ws = wb["登録番号"]
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            corporate_number = self.generate_corporate_number(ws.cell(row_num, Columns.invoice_number).value)
            address = fetched_data.get(corporate_number, {}).get("address", "")
            registered_name = fetched_data.get(corporate_number, {}).get("name", "")

            ws.cell(row_num, Columns.corporate_number).value = corporate_number
            ws.cell(row_num, Columns.address).value = address
            ws.cell(row_num, Columns.tax_agency_registered_name).value = registered_name

        wb.save(self.target_excel)
        wb.close()
